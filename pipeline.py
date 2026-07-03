"""
pipeline.py — the FMCG reporting engine (demo, dummy data).

What it does, in one run:
  1. COLLECT   three distributor files in three different formats
  2. CLEAN     junk headers, duplicates, typo'd SKU names, mixed date
               formats, missing values, stray total rows
  3. REPORT    a formatted Excel report: sales vs target by region,
               top SKUs, distributor summary, data-quality log
  4. DELIVER   (on Windows) email the report via Outlook — guarded so
               the pipeline also runs anywhere without Outlook

Run: python pipeline.py
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

BASE = Path(__file__).parent
RAW, OUT = BASE / "raw", BASE / "output"
OUT.mkdir(exist_ok=True)

quality_log = []   # every fix gets logged — this becomes the report's audit trail
def log(source, issue, action, count):
    quality_log.append({"Source file": source, "Issue found": issue,
                        "Action taken": action, "Rows affected": count})

# --------------------------------------------------------------------------
# Canonical product catalogue — one true name per SKU
# --------------------------------------------------------------------------
CATALOGUE = {
    "SAV-TEA-050": "Savanna Chai 50g",
    "SAV-TEA-100": "Savanna Chai 100g",
    "SAV-TEA-250": "Savanna Chai 250g",
    "SAV-COF-100": "Savanna Coffee 100g",
    "SAV-COF-250": "Savanna Coffee 250g",
    "SAV-COC-500": "Savanna Cocoa 500g",
    "SAV-JUI-1LT": "Savanna Juice 1L",
    "SAV-JUI-500": "Savanna Juice 500ml",
}
NAME_TO_CODE = {v.lower(): k for k, v in CATALOGUE.items()}

def normalise_sku(value):
    """Map any SKU representation — code, clean name, or typo'd name — to its code."""
    v = str(value).strip()
    if v.upper() in CATALOGUE:
        return v.upper()
    key = re.sub(r"\s+", " ", v.lower())
    key = (key.replace("savana", "savanna").replace("sav ", "savanna ")
              .replace("100gm", "100g").replace("50 g", "50g")
              .replace("1 litre", "1l").replace("1lt", "1l"))
    return NAME_TO_CODE.get(key)          # None if unrecognisable

# --------------------------------------------------------------------------
# 1. COLLECT + 2. CLEAN — one loader per source, all ending in one schema:
#    Date | Distributor | Region | SKU | Units | UnitPrice | Outlets
# --------------------------------------------------------------------------
def load_pwani():
    f = "pwani_daily_sales_june.csv"
    df = pd.read_csv(RAW / f)
    before = len(df)
    df = df.drop_duplicates()
    log(f, "Duplicate rows", "Removed", before - len(df))
    zero = (df["Units"] <= 0).sum()
    df = df[df["Units"] > 0]
    log(f, "Zero/negative unit rows", "Removed", int(zero))
    df["Date"] = pd.to_datetime(df["Date"], format="%Y-%m-%d")
    return df

def load_highlands():
    f = "highlands_june_2026.xlsx"
    df = pd.read_excel(RAW / f, skiprows=3)          # junk header rows
    log(f, "Junk header rows above data", "Skipped", 3)
    total_rows = df["Date"].astype(str).str.upper().str.contains("TOTAL").sum()
    df = df[~df["Date"].astype(str).str.upper().str.contains("TOTAL")]
    log(f, "Stray TOTAL row", "Removed", int(total_rows))

    missing = df["Units"].isna().sum()
    df = df.dropna(subset=["Units"])
    log(f, "Missing unit values", "Removed & flagged", int(missing))

    # mixed date formats: 15/06/2026 and 3-6-26 in the same column
    def parse_date(v):
        v = str(v).strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%y"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                continue
        return pd.NaT
    df["Date"] = df["Date"].apply(parse_date)
    log(f, "Mixed date formats", "Standardised", len(df))

    before = len(df)
    df["SKU"] = df["SKU"].apply(normalise_sku)
    unmatched = df["SKU"].isna().sum()
    df = df.dropna(subset=["SKU"])
    log(f, "Typo'd/inconsistent SKU names", "Mapped to catalogue codes", before - int(unmatched))
    if unmatched:
        log(f, "Unrecognisable SKU names", "Removed & flagged", int(unmatched))
    return df

def load_lakeview():
    f = "lakeview_export_jun26.csv"
    df = pd.read_csv(RAW / f)
    df = df.rename(columns={                          # different schema entirely
        "txn_date": "Date", "product_code": "SKU", "qty_sold": "Units",
        "price_per_unit": "UnitPrice", "sales_region": "Region",
        "dist_name": "Distributor", "outlet_count": "Outlets",
    })
    log(f, "Non-standard column names/order", "Mapped to standard schema", len(df))
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y")
    return df

frames = [load_pwani(), load_highlands(), load_lakeview()]
data = pd.concat(frames, ignore_index=True)
data["SKU"] = data["SKU"].apply(normalise_sku)
data["Product"] = data["SKU"].map(CATALOGUE)
data["Revenue"] = data["Units"] * data["UnitPrice"]

targets = pd.read_csv(RAW / "june_targets.csv")

# --------------------------------------------------------------------------
# 3. REPORT — formatted Excel workbook
# --------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT, GOOD, BAD, HEADFILL = "0F4C5C", "1E7D32", "C62828", "E8EEF0"
thin = Side(style="thin", color="CCCCCC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=ACCENT)
        cell.alignment = Alignment(horizontal="center")
        cell.border = box

def title(ws, text, ncols):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=14, name="Arial", color=ACCENT)
    ws["A2"] = f"Savanna Beverages (demo data) · June 2026 · generated {datetime.now():%d %b %Y %H:%M}"
    ws["A2"].font = Font(size=9, name="Arial", color="777777")

def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 1: Sales vs Target ---
ws = wb.active
ws.title = "Sales vs Target"
title(ws, "SALES vs TARGET — BY REGION", 5)
region = (data.groupby("Region", as_index=False)
              .agg(Revenue=("Revenue", "sum"), Units=("Units", "sum"))
              .merge(targets, on="Region", how="left"))
headers = ["Region", "Units Sold", "Revenue (Ksh)", "Target (Ksh)", "Achievement %"]
ws.append([]); ws.append(headers)
style_header(ws, 4, len(headers))
r0 = 5
for i, row in region.iterrows():
    r = r0 + i
    ws.cell(row=r, column=1, value=row["Region"])
    ws.cell(row=r, column=2, value=int(row["Units"]))
    ws.cell(row=r, column=3, value=float(row["Revenue"]))
    ws.cell(row=r, column=4, value=float(row["MonthlyTargetKsh"]))
    ws.cell(row=r, column=5, value=f"=C{r}/D{r}")
    ws.cell(row=r, column=2).number_format = "#,##0"
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4).number_format = "#,##0"
    ws.cell(row=r, column=5).number_format = "0.0%"
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = box
        ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
last = r0 + len(region) - 1
tr = last + 1
ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=10)
for col, letter in [(2, "B"), (3, "C"), (4, "D")]:
    ws.cell(row=tr, column=col, value=f"=SUM({letter}{r0}:{letter}{last})")
    ws.cell(row=tr, column=col).number_format = "#,##0"
    ws.cell(row=tr, column=col).font = Font(bold=True, name="Arial", size=10)
ws.cell(row=tr, column=5, value=f"=C{tr}/D{tr}")
ws.cell(row=tr, column=5).number_format = "0.0%"
ws.cell(row=tr, column=5).font = Font(bold=True, name="Arial", size=10)
for c in range(1, 6):
    ws.cell(row=tr, column=c).fill = PatternFill("solid", start_color=HEADFILL)
    ws.cell(row=tr, column=c).border = box
autofit(ws, [16, 12, 16, 16, 15])

# --- Sheet 2: Top SKUs ---
ws2 = wb.create_sheet("Top SKUs")
title(ws2, "PRODUCT PERFORMANCE — JUNE", 4)
sku = (data.groupby(["SKU", "Product"], as_index=False)
           .agg(Units=("Units", "sum"), Revenue=("Revenue", "sum"))
           .sort_values("Revenue", ascending=False))
headers2 = ["SKU", "Product", "Units", "Revenue (Ksh)"]
ws2.append([]); ws2.append(headers2)
style_header(ws2, 4, len(headers2))
for i, row in sku.reset_index(drop=True).iterrows():
    r = 5 + i
    ws2.cell(row=r, column=1, value=row["SKU"])
    ws2.cell(row=r, column=2, value=row["Product"])
    ws2.cell(row=r, column=3, value=int(row["Units"]))
    ws2.cell(row=r, column=4, value=float(row["Revenue"]))
    ws2.cell(row=r, column=3).number_format = "#,##0"
    ws2.cell(row=r, column=4).number_format = "#,##0"
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = box
        ws2.cell(row=r, column=c).font = Font(name="Arial", size=10)
autofit(ws2, [14, 24, 10, 16])

# --- Sheet 3: By Distributor ---
ws3 = wb.create_sheet("By Distributor")
title(ws3, "DISTRIBUTOR SUMMARY — JUNE", 4)
dist = (data.groupby("Distributor", as_index=False)
            .agg(Units=("Units", "sum"), Revenue=("Revenue", "sum"),
                 Outlets=("Outlets", "sum")))
headers3 = ["Distributor", "Units", "Revenue (Ksh)", "Outlet Touches"]
ws3.append([]); ws3.append(headers3)
style_header(ws3, 4, len(headers3))
for i, row in dist.iterrows():
    r = 5 + i
    ws3.cell(row=r, column=1, value=row["Distributor"])
    ws3.cell(row=r, column=2, value=int(row["Units"]))
    ws3.cell(row=r, column=3, value=float(row["Revenue"]))
    ws3.cell(row=r, column=4, value=int(row["Outlets"]))
    for c, fmt in [(2, "#,##0"), (3, "#,##0"), (4, "#,##0")]:
        ws3.cell(row=r, column=c).number_format = fmt
    for c in range(1, 5):
        ws3.cell(row=r, column=c).border = box
        ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
autofit(ws3, [22, 10, 16, 14])

# --- Sheet 4: Data Quality Log (the differentiator) ---
ws4 = wb.create_sheet("Data Quality Log")
title(ws4, "DATA QUALITY — ISSUES CAUGHT & FIXED THIS RUN", 4)
headers4 = ["Source file", "Issue found", "Action taken", "Rows affected"]
ws4.append([]); ws4.append(headers4)
style_header(ws4, 4, len(headers4))
for i, entry in enumerate(quality_log):
    r = 5 + i
    ws4.cell(row=r, column=1, value=entry["Source file"])
    ws4.cell(row=r, column=2, value=entry["Issue found"])
    ws4.cell(row=r, column=3, value=entry["Action taken"])
    ws4.cell(row=r, column=4, value=entry["Rows affected"])
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = box
        ws4.cell(row=r, column=c).font = Font(name="Arial", size=10)
autofit(ws4, [26, 32, 30, 14])

report_path = OUT / "Savanna_Monthly_Sales_Report_June2026.xlsx"
wb.save(report_path)
print(f"Report written: {report_path}")

# --------------------------------------------------------------------------
# 4. DELIVER — Outlook dispatch (Windows only; skipped gracefully elsewhere)
# --------------------------------------------------------------------------
SEND_EMAIL = False        # set True on your Windows machine to actually send
RECIPIENT = "manager@example.com"

ach = region.merge(targets, on="Region", how="left", suffixes=("", "_t"))
total_rev = data["Revenue"].sum()
total_target = targets["MonthlyTargetKsh"].sum()
summary_line = (f"June revenue Ksh {total_rev:,.0f} vs target Ksh {total_target:,.0f} "
                f"({total_rev/total_target:.1%} achievement). "
                f"Top product: {sku.iloc[0]['Product']}.")

if SEND_EMAIL and sys.platform == "win32":
    import win32com.client
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = RECIPIENT
    mail.Subject = f"Monthly Sales Report — June 2026 (auto-generated)"
    mail.Body = ("Good morning,\n\nPlease find attached the June sales report.\n\n"
                 + summary_line +
                 "\n\nThis report was compiled and sent automatically.\n")
    mail.Attachments.Add(str(report_path.resolve()))
    mail.Send()
    print(f"Emailed to {RECIPIENT} via Outlook.")
else:
    print("Email step skipped (SEND_EMAIL=False or not on Windows).")
    print("Summary that would be sent:", summary_line)
